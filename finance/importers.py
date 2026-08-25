"""
Workday journal export -> :class:`WorkdayTransaction` rows.

Reads CSV and .xlsx alike: Workday will hand you either, and which button
someone happened to press should not decide whether the ledger can read the
file. :func:`read_table` reduces both to the same rows-of-cells shape, and
nothing downstream of it knows which format it came from.

The importer is deliberately forgiving about formatting, because Workday's
export dialect drifts: it emits a UTF-8 BOM, wraps negatives in parentheses,
includes thousands separators and currency symbols, embeds newlines inside
quoted memos, pads headers with non-breaking spaces, and puts a report title
above the table when exporting to Excel.

Identity is the interesting part; see :func:`_resolve_duplicates`.
"""
import csv
import datetime
import io
import re
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import models, transaction
from django.utils import timezone
from django.utils.crypto import get_random_string

from finance.models import (MEMO_SEPARATOR, WorkdayTransaction, _identity_text,
                            column_aliases)

# Canonical Workday columns. Keys are the normalised header, values are the
# aliases we have actually seen come out of Workday.
COLUMN_ALIASES = {
    'accounting_date': ('accounting date',),
    'debit_amount': ('debit amount',),                    # ignored, per spec
    'credit_amount': ('credit amount',),                  # ignored, per spec
    'credit_minus_debit': ('credit minus debit', 'credit - debit', 'credit minus debit amount'),
    'operational_transaction': ('operational transaction', 'operational transaction id'),
    'supplier': ('supplier',),
    'employee': ('employee',),
    'journal': ('journal',),
    'journal_line_memo': ('journal line memo', 'line memo'),
    'header_memo': ('header memo',),
    'fund': ('fund',),
    'cost_center': ('cost center', 'cost centre'),
    'ledger_account': ('ledger account',),
    'spend_category': ('spend category',),
    'revenue_category': ('revenue category',),
    'activity': ('activity',),
    'student_organization': ('student organization', 'student org'),
    'program': ('program',),
}

# Columns explicitly discarded: direction is derived solely from Credit Minus Debit.
IGNORED_COLUMNS = ('debit_amount', 'credit_amount')

# Columns that become dedicated model fields rather than worktags.
STRUCTURAL_COLUMNS = (
    'accounting_date', 'credit_minus_debit', 'operational_transaction',
    'supplier', 'employee', 'journal_line_memo', 'header_memo',
) + IGNORED_COLUMNS

# Columns whose *presence* proves this is a Workday journal export. Only the
# date and the amount are required to have a value on every line: journal entry
# lines legitimately carry no Operational Transaction.
REQUIRED_COLUMNS = ('accounting_date', 'credit_minus_debit', 'operational_transaction')

_DATE_FORMATS = ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%d/%m/%Y', '%b %d, %Y', '%m-%d-%Y')


# How far down a sheet to look for the header row. Workday's "export to Excel"
# puts a report title and a filter summary above the table; a CSV usually
# starts straight in, but the same scan costs nothing there.
HEADER_SCAN_ROWS = 25

XLSX_EXTENSIONS = ('.xlsx', '.xlsm')
CSV_EXTENSIONS = ('.csv', '.tsv', '.txt')
# Every xlsx is a zip, and every zip starts 'PK'. Extensions lie; this does not.
ZIP_MAGIC = b'PK\x03\x04'


class ImportError_(Exception):
    """ Raised when the file cannot be parsed at all (bad headers, not a CSV). """


def _text(value):
    """
    One cell as trimmed text, whatever type it arrived as.

    A CSV yields strings; a spreadsheet yields numbers, dates and ``None``.
    Everything downstream wants text, so the difference stops here.
    """
    if value is None:
        return ''
    if isinstance(value, str):
        return value.replace('\xa0', ' ').strip()
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # Excel stores every number as a float, so an id would read '25070087.0'.
        return str(int(value))
    return str(value).strip()


def _is_filled(value):
    """
    Whether a cell holds anything once whitespace is discounted.

    Used when scoring candidate header rows, where a row of empty strings
    and a row of spaces have to count the same.
    """
    return bool(_text(value))


def _clean_header(raw):
    """
    A header reduced to lowercase words: ``"  Credit - Debit "`` -> ``"credit debit"``.

    Both sides of every alias comparison go through this, which is the point of
    having it as a function. They did not always: the incoming header was
    cleaned and the alias table was not, so an alias written with any
    punctuation could never match. ``'credit - debit'`` sat in
    :data:`COLUMN_ALIASES` matching nothing, and an export whose amount column
    was headed that way was rejected as missing a required column -- a file
    that looks fine in Excel, refused for a reason nobody could see.
    """
    cleaned = raw.replace('﻿', '').replace('\xa0', ' ').strip().lower()
    return re.sub(r'[^a-z0-9]+', ' ', cleaned).strip()


#: ``{cleaned alias: canonical name}``, flattened once from
#: :data:`COLUMN_ALIASES`. Building it here rather than scanning the nested
#: table per header means a spelling cannot be added in a form that never
#: matches: it is normalised on the way in, exactly like the header it is
#: compared against.
_ALIAS_LOOKUP = {
    _clean_header(alias): canonical
    for canonical, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}


def normalise_header(raw):
    """
    ``"  Credit Minus Debit "`` -> ``"credit_minus_debit"``.

    Aliases added in the admin are consulted first, so a column Workday has
    relabelled can be taught to the importer without a deploy.
    """
    if raw is None:
        return ''
    cleaned = _clean_header(raw)

    from_admin = column_aliases().get(cleaned)
    if from_admin:
        return from_admin
    canonical = _ALIAS_LOOKUP.get(cleaned)
    if canonical:
        return canonical
    return re.sub(r'\s+', '_', cleaned)


def parse_decimal(raw):
    """
    Parse Workday's money formatting into a Decimal.

    Handles ``$1,200.00``, ``(1,200.00)`` (negative), ``1200``, ``-1,200.00``,
    ``1,200.00 USD`` and empty cells -- and, from a spreadsheet, a number that
    arrives already typed.
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, Decimal)):
        return Decimal(raw).quantize(Decimal('0.01'))
    if isinstance(raw, float):
        # via str() so 17.91 does not become 17.910000000000000142.
        return Decimal(str(raw)).quantize(Decimal('0.01'))
    text = str(raw).replace('\xa0', ' ').strip()
    if not text:
        return None
    negative = text.startswith('(') and text.endswith(')')
    if negative:
        text = text[1:-1]
    text = re.sub(r'[^0-9.\-]', '', text)
    if not text or text in ('-', '.', '-.'):
        return None
    try:
        value = Decimal(text)
    except InvalidOperation:
        return None
    if negative:
        value = -value
    return value.quantize(Decimal('0.01'))


def parse_date(raw):
    """ A date from Workday's text formats, or from an already-typed cell. """
    if raw is None:
        return None
    # A spreadsheet stores dates as numbers and openpyxl hands them back as
    # datetimes, so there is nothing to parse.
    if isinstance(raw, datetime.datetime):
        return raw.date()
    if isinstance(raw, datetime.date):
        return raw
    text = str(raw).replace('\xa0', ' ').strip()
    if not text:
        return None
    text = text.split(' ')[0] if ' ' in text and ',' not in text else text
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def build_memo(row):
    """ Journal Line Memo + Header Memo, concatenated, de-duplicated. """
    parts = []
    for key in ('journal_line_memo', 'header_memo'):
        value = _text(row.get(key))
        if value and value not in parts:
            parts.append(value)
    return MEMO_SEPARATOR.join(parts)


class RowResult(object):
    """ Outcome of a single CSV row. """
    CREATED, DUPLICATE, ERROR = 'created', 'duplicate', 'error'

    def __init__(self, line_number, status, message='', transaction=None, preview=None,
                 warning=''):
        """
        Record what happened to one row.

        ``transaction`` is set only on a successful create. ``preview`` holds
        the parsed field values so a dry run can show the operator what *would*
        be written without writing it. ``warning`` is set on a row that is
        importing but looks like something already on file -- see
        :func:`_find_near_duplicates`.
        """
        self.line_number = line_number
        self.status = status
        self.message = message
        self.transaction = transaction
        self.preview = preview or {}
        self.warning = warning

    @property
    def is_error(self):
        """ Whether this row failed outright, as opposed to being a duplicate. """
        return self.status == self.ERROR


class ImportResult(object):
    """
    Every row of one import, and the tallies the UI reports from them.

    The same object is produced by a dry run and by a real import, which is
    what lets the confirmation page state an exact row count before anything
    is committed. Rows are kept individually rather than only counted so an
    error can be reported against the line it came from.
    """

    def __init__(self, filename=''):
        """ Start an empty tally for one uploaded file. """
        self.filename = filename
        self.rows = []
        self.headers = []
        self.unmapped_headers = []

    def add(self, row_result):
        """ Record the outcome of one more row. """
        self.rows.append(row_result)

    def _count(self, status):
        """ How many rows ended in ``status``. """
        return sum(1 for r in self.rows if r.status == status)

    @property
    def created(self):
        """ The rows that produced a transaction. """
        return [r for r in self.rows if r.status == RowResult.CREATED]

    @property
    def created_count(self):
        """ How many new transactions this import would write, or wrote. """
        return self._count(RowResult.CREATED)

    @property
    def duplicate_count(self):
        """
        How many rows were already in the ledger.

        A high count is normal, not a warning: exports overlap, and matching
        on the row fingerprint is exactly what stops that double-counting.
        """
        return self._count(RowResult.DUPLICATE)

    @property
    def error_count(self):
        """ How many rows could not be parsed at all. """
        return self._count(RowResult.ERROR)

    @property
    def errors(self):
        """ The failed rows, each still carrying its line number and message. """
        return [r for r in self.rows if r.is_error]

    @property
    def suspects(self):
        """
        New rows that look like a line the ledger already holds.

        These are importing, not blocked -- the fingerprint says they are new
        and the fingerprint is what the ledger trusts. They are surfaced
        because the one way this importer can double-count is a line whose
        exported detail changed between two exports of the same charge, and
        that is invisible in a count of "new rows".
        """
        return [r for r in self.rows if r.warning]

    @property
    def total(self):
        """ Every data row seen, whatever became of it. """
        return len(self.rows)

    @property
    def ok(self):
        """
        Whether the file parsed cleanly end to end.

        Duplicates do not count against it -- they are the expected case.
        """
        return self.error_count == 0

    def summary(self):
        """ One-line tally for the confirmation page and the messages framework. """
        return "%s imported, %s skipped as duplicates, %s errors" % (
            self.created_count, self.duplicate_count, self.error_count)

    def unmapped_spend_categories(self):
        """
        Workday spend categories on the new lines that no rule accounts for.

        Each one is a category the Treasurer would otherwise pick by hand on
        every line carrying it, forever. Reported after an import because that
        is the moment a new one appears, and one row in the admin settles it.
        """
        from finance.suggestions import unmapped_spend_categories
        return unmapped_spend_categories(
            [r.transaction for r in self.created if r.transaction is not None
             and r.transaction.net_amount < 0])


def _resolve_duplicates(candidates):
    """
    Decide which parsed rows are new and which the ledger already holds.

    Workday gives us nothing that identifies a line. ``Operational Transaction``
    names the *document* -- one supplier invoice covers a dozen lines, and a
    journal entry has none at all -- and every other column can legitimately
    repeat, because two identical charges in a month really are two charges.

    So identity is "the whole exported line", and the ambiguity that leaves is
    settled by counting rather than by guessing:

        the ledger should hold as many copies of a line as the fullest
        export has ever shown

    A file listing the Spotify subscription twice imports both. Re-uploading it
    imports neither, because two are already on file. A later export covering
    the same period plus a third charge imports exactly the third. The
    arithmetic is per file, so ordering within the file does not matter.

    The known limit: an export deliberately narrowed to a single line cannot
    add a second copy of a charge already on file -- it is indistinguishable
    from a re-upload of one line. Those land in the queue as duplicates with
    the count spelled out, and an encumbrance can carry the odd genuine case.

    :param candidates: ``(line_number, transaction)`` in file order.
    :returns: ``(to_create, decisions)`` where decisions maps line number to
        either ``None`` (create) or a message explaining the skip.
    """
    fingerprints = {txn.row_fingerprint for _, txn in candidates}
    already_held = defaultdict(int)
    highest_ordinal = defaultdict(int)
    if fingerprints:
        # Count and high-water mark are deliberately two different numbers.
        # How many copies are on file decides what counts as a duplicate; the
        # highest occurrence number already taken decides what a new row may be
        # numbered. They diverge the moment hard_delete() removes anything but
        # the last occurrence, and numbering from the count would then collide
        # with a surviving row and abort the whole import on the unique
        # constraint.
        held = (WorkdayTransaction.objects
                .filter(row_fingerprint__in=fingerprints)
                .values('row_fingerprint')
                .annotate(n=models.Count('pk'),
                          top=models.Max('fingerprint_ordinal')))
        for row in held:
            already_held[row['row_fingerprint']] = row['n']
            highest_ordinal[row['row_fingerprint']] = row['top'] or 0

    seen_in_file = Counter()
    to_create, decisions = [], {}

    for line_number, txn in candidates:
        seen_in_file[txn.row_fingerprint] += 1
        occurrence = seen_in_file[txn.row_fingerprint]
        held = already_held[txn.row_fingerprint]

        if occurrence <= held:
            decisions[line_number] = (
                "Already imported." if held == 1 else
                "Already imported — the ledger holds %s identical lines and this file "
                "shows %s." % (held, occurrence))
            continue

        # Occurrences 1..held are on file; this one continues the sequence.
        highest_ordinal[txn.row_fingerprint] += 1
        txn.fingerprint_ordinal = highest_ordinal[txn.row_fingerprint]
        to_create.append(txn)
        decisions[line_number] = None

    return to_create, decisions


def _near_duplicate_key(txn):
    """
    The part of a line that a re-parse cannot plausibly change.

    Date, amount and who the document was with come off the export as their own
    columns and survive any reasonable disagreement about quoting, column
    naming or memo joining. The memo and the worktags do not, which is exactly
    why they are the fields that go wrong quietly.

    ``None`` when the line names neither a document nor a person: journal entry
    lines carry only a date and an amount, and two of those matching means
    nothing -- one October journal in this ledger holds four separate $100
    projector rentals.
    """
    document = _identity_text(txn.operational_transaction)
    supplier = _identity_text(txn.supplier)
    employee = _identity_text(txn.employee)
    if not (document or supplier or employee):
        return None
    return (txn.accounting_date,
            Decimal(txn.net_amount or 0).quantize(Decimal('0.01')),
            document, supplier, employee)


def _find_near_duplicates(to_create):
    """
    Rows that are new by fingerprint but look like a line already on file.

    The fingerprint is an exact hash of the whole exported line, which is the
    only honest answer to "is this the same line" -- but it means any drift in
    how a line is *read* reads as a brand new charge. A bad CSV dialect that
    shifted a memo across four worktag columns did exactly that: the line
    imported a second time, months after the first copy had been reconciled,
    and nothing in the import counts looked unusual.

    So the fingerprint still decides what gets written, and this decides what
    gets said about it. A match here is not proof of a duplicate -- one
    supplier invoice can legitimately carry two lines of the same amount -- so
    these rows import and are flagged, never blocked.

    :param to_create: ``(line_number, transaction)`` for the rows being written.
    :returns: ``{line_number: [existing WorkdayTransaction, ...]}``
    """
    keyed = defaultdict(list)
    for line_number, txn in to_create:
        key = _near_duplicate_key(txn)
        if key is not None:
            keyed[key].append((line_number, txn))
    if not keyed:
        return {}

    # Narrowed in SQL by the two columns that are indexed and selective, then
    # matched exactly in Python: date and amount together are cheap to filter
    # on and leave few enough rows to compare properly.
    candidates = WorkdayTransaction.objects.filter(
        accounting_date__in={key[0] for key in keyed},
        net_amount__in={key[1] for key in keyed})

    matches = {}
    for existing in candidates:
        key = _near_duplicate_key(existing)
        if key not in keyed:
            continue
        for line_number, txn in keyed[key]:
            # Same fingerprint is the deliberate case, not the suspect one:
            # a second copy of an identical charge, numbered by its ordinal.
            if existing.row_fingerprint == txn.row_fingerprint:
                continue
            matches.setdefault(line_number, []).append(existing)
    return matches


def _describe_near_duplicates(existing):
    """
    The one-line warning for a row :func:`_find_near_duplicates` flagged.

    Names the rows it resembles so the Treasurer can open them, and says what
    it means, because "possible duplicate" on its own invites either ignoring
    every one of them or deleting a legitimate second line.
    """
    if not existing:
        return ''
    which = ", ".join("#%s" % txn.pk for txn in existing[:3])
    if len(existing) > 3:
        which += " and %s more" % (len(existing) - 3)
    return ("Same date, amount, document and payee as %s already in the ledger, but the "
            "memo or worktags differ, so this imports as a separate line. Check %s before "
            "confirming: if it is the same charge read differently, this would double-count "
            "it." % (which, "them" if len(existing) > 1 else "it"))


def _decode(file_obj):
    """ Read an uploaded file into text, tolerating BOMs and cp1252 exports. """
    raw = file_obj.read()
    if isinstance(raw, str):
        return raw
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8', errors='replace')


def _looks_like_header(cells):
    """ True if this row names every column the importer cannot do without. """
    found = {normalise_header(_text(cell)) for cell in cells if _is_filled(cell)}
    return all(column in found for column in REQUIRED_COLUMNS)


def _find_header(rows):
    """
    Pick the header row out of the first :data:`HEADER_SCAN_ROWS` rows.

    Workday's Excel export puts a report title and the filters used above the
    table, so the header is rarely the first row. Returns
    ``(header_cells, [(line number, row), ...])`` with line numbers counted
    from 1 as a spreadsheet or a text editor would show them.
    """
    for index, cells in enumerate(rows[:HEADER_SCAN_ROWS]):
        if _looks_like_header(cells):
            return cells, list(enumerate(rows[index + 1:], start=index + 2))
    if not rows:
        raise ImportError_("That file has no rows in it.")

    # Nothing matched: report against the most table-looking row we saw, which
    # gives a far more useful message than "row 1 was a title".
    widest = max(rows[:HEADER_SCAN_ROWS], key=lambda cells: sum(1 for c in cells if _is_filled(c)))
    found = ", ".join(_text(c) for c in widest if _is_filled(c))
    raise ImportError_(
        "This doesn't look like a Workday journal export — no row in the first %s "
        "names Accounting Date, Credit Minus Debit and Operational Transaction. "
        "The most likely header row reads: %s" % (HEADER_SCAN_ROWS, found or "(nothing)"))


def _read_csv(raw):
    """
    Parse a CSV or tab-separated export into a list of rows.

    Returns raw cell lists, not dictionaries: the header row is not
    necessarily the first one, so working out which row is the header is a
    later step (see :func:`find_header_row`).
    """
    text = _decode(io.BytesIO(raw) if isinstance(raw, bytes) else raw)
    if not text.strip():
        raise ImportError_("That file is empty.")

    # Sniff the *delimiter* only -- Workday sometimes hands out tab-separated
    # files -- and keep Excel's quoting rules regardless.
    #
    # Taking the sniffer's whole dialect was silently corrupting data: on the
    # FY26 export it guessed doublequote=False, which breaks the RFC-4180 ""
    # escape, so a memo reading  Planar 22" touchscreen  came back with a stray
    # quote on the end. Two lines in 314, and nothing on screen to suggest it.
    try:
        delimiter = csv.Sniffer().sniff(text[:8192], delimiters=',\t;').delimiter
    except csv.Error:
        delimiter = ','
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


def _read_xlsx(raw):
    """
    Every row of the first worksheet that holds a Workday table.

    A workbook can carry more than one sheet -- a summary tab, or last term's
    export left behind -- so the sheets are searched rather than assuming the
    first one, and the first that contains the required columns wins.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:                                     # pragma: no cover
        raise ImportError_(
            "Reading .xlsx files needs the openpyxl package, which is not installed on "
            "this server. Export the journal as CSV instead, or ask the webmaster to "
            "install it.")

    try:
        # read_only streams the sheet instead of building the whole object
        # graph; data_only takes the cached result of a formula rather than
        # the formula text, which is what a Workday export contains.
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportError_("That .xlsx file could not be opened (%s). If Excel can open "
                           "it, try re-saving it as .xlsx or CSV." % exc)

    try:
        best = None
        for sheet in workbook.worksheets:
            # Ignore the extent the file declares for itself and work it out
            # from the cells that are actually there.
            #
            # A read-only worksheet trusts the ``<dimension>`` element in the
            # sheet XML and clips every row to that width. Workday's exporter
            # writes one that understates the table -- ``A1:A1`` on a
            # thirty-column export -- so every row arrived one cell wide and
            # the header read simply ``Accounting Date``. The file opened, the
            # sheet was found, and the only symptom was the importer insisting
            # a perfectly good export was not one.
            #
            # openpyxl provides this for exactly that case, and it costs
            # nothing on a file whose dimension was honest.
            if hasattr(sheet, 'reset_dimensions'):
                sheet.reset_dimensions()
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if any(_looks_like_header(cells) for cells in rows[:HEADER_SCAN_ROWS]):
                return rows
            if best is None and any(any(_is_filled(c) for c in cells) for cells in rows):
                best = rows
        if best is None:
            raise ImportError_("That workbook has no data in it.")
        # No sheet matched; hand back the first non-empty one so _find_header
        # can produce its more specific complaint.
        return best
    finally:
        workbook.close()


def read_table(file_obj, filename=''):
    """
    Reduce an upload to ``(header cells, [(line number, row cells), ...])``.

    The only place that knows or cares which format arrived. Dispatch is on the
    file's own bytes first and its name second, because browsers and operating
    systems both get content types wrong, and a spreadsheet saved as ``.csv``
    is a genuinely common mistake.
    """
    raw = file_obj.read()
    if isinstance(raw, str):
        raw = raw.encode('utf-8')

    name = (filename or getattr(file_obj, 'name', '') or '').lower()
    if raw[:4] == ZIP_MAGIC:
        rows = _read_xlsx(raw)
    elif name.endswith(XLSX_EXTENSIONS):
        # Named .xlsx but not a zip: almost always an .xls or a CSV renamed.
        raise ImportError_(
            "That file is named .xlsx but is not a valid Excel workbook. If it came from "
            "an older Excel, open it and 'Save As' .xlsx or CSV.")
    else:
        rows = _read_csv(raw)

    return _find_header(rows)


def import_workday_export(file_obj, user=None, filename='', dry_run=False):
    """
    Parse a Workday export and create :class:`WorkdayTransaction` rows.

    Takes either a CSV or an .xlsx workbook -- Workday will hand you either,
    and which one you got should not decide whether the ledger can read it.
    See :func:`read_table` for how the two are reduced to the same shape.

    Rows already on file are skipped, never updated -- the bank truth is
    immutable, so a re-import is a no-op rather than a rewrite, and uploading
    the same file twice is safe. :func:`_resolve_duplicates` explains how
    "already on file" is decided.

    :param dry_run: parse and validate without writing anything.
    :returns: :class:`ImportResult`
    """
    header_row, data_rows = read_table(file_obj, filename)

    headers = [normalise_header(h) for h in header_row]
    result = ImportResult(filename=filename)
    result.headers = headers

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        pretty = ", ".join(c.replace('_', ' ').title() for c in missing)
        raise ImportError_(
            "This doesn't look like a Workday journal export — missing required column(s): %s. "
            "Found: %s" % (pretty, ", ".join(h.replace('_', ' ').title() for h in headers if h)))

    known = set(COLUMN_ALIASES.keys())
    result.unmapped_headers = [h for h in headers if h and h not in known]

    candidates = []
    previews = {}

    for line_number, raw_row in data_rows:
        if not any(_is_filled(cell) for cell in raw_row):
            continue  # blank spacer row

        # A row wider than the header means the line did not split where the
        # header did, and every cell after the break is under the wrong
        # column. Silently truncating to the header width -- which is what
        # indexing alone does -- turns that into a row that imports cleanly,
        # reads plausibly, and is wrong in every worktag. It has happened: a
        # memo reading  Stereo 1/4" cables, Maintenance and Repair  was split
        # at its commas by a bad CSV dialect, shifting the fund, cost center,
        # ledger account and spend category one column each and pushing the
        # student organization off the end. Nothing downstream can tell that
        # from a real line, and its fingerprint will never match the same line
        # exported again, so it double-counts on the next import.
        #
        # Trailing empty cells are not that: a stray delimiter at the end of a
        # line is ordinary, so only cells with something in them count.
        overflow = [cell for cell in raw_row[len(headers):] if _is_filled(cell)]
        if overflow:
            result.add(RowResult(
                line_number, RowResult.ERROR,
                "This line has %s more value(s) than there are columns, so the "
                "columns after the overflow cannot be trusted. Usually a quote or a "
                "delimiter inside a memo. Extra value(s): %s"
                % (len(overflow), ", ".join(repr(_text(c)) for c in overflow[:3]))))
            continue

        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = raw_row[index] if index < len(raw_row) else ''

        # Blank on every journal entry line, which is a normal thing for an
        # export to contain. Identity does not depend on it.
        op_txn = _text(row.get('operational_transaction'))

        net_amount = parse_decimal(row.get('credit_minus_debit'))
        if net_amount is None:
            result.add(RowResult(line_number, RowResult.ERROR,
                                 "Could not read 'Credit Minus Debit' (got %r)."
                                 % _text(row.get('credit_minus_debit'))))
            continue
        if net_amount == 0:
            result.add(RowResult(line_number, RowResult.ERROR,
                                 "Net amount is $0.00 — nothing to reconcile."))
            continue

        accounting_date = parse_date(row.get('accounting_date'))
        if accounting_date is None:
            result.add(RowResult(line_number, RowResult.ERROR,
                                 "Could not read 'Accounting Date' (got %r)."
                                 % _text(row.get('accounting_date'))))
            continue

        worktags = {}
        for header, value in row.items():
            if header in STRUCTURAL_COLUMNS or not header:
                continue
            cleaned = _text(value)
            if cleaned:
                worktags[header] = cleaned

        # Kept separately as well as concatenated into `memo`, because it is
        # what seeds an allocation's Description and the two memos cannot be
        # pulled back apart reliably once joined.
        line_memo = _text(row.get('journal_line_memo'))
        if line_memo:
            worktags['journal_line_memo'] = line_memo

        txn = WorkdayTransaction(
            operational_transaction=op_txn,
            accounting_date=accounting_date,
            net_amount=net_amount,
            supplier=_text(row.get('supplier'))[:255],
            employee=_text(row.get('employee'))[:255],
            memo=build_memo(row),
            worktags_json=worktags,
            imported_by=user,
            source_file=filename[:255],
        )
        txn.row_fingerprint = txn.compute_fingerprint()
        candidates.append((line_number, txn))
        previews[line_number] = {
            'operational_transaction': op_txn,
            'reference': txn.reference,
            'accounting_date': accounting_date,
            'net_amount': net_amount,
            'payee': txn.payee,
            'ledger_account': worktags.get('ledger_account', ''),
            'is_projection': txn.defaults_to_projection,
        }

    # Deferred to here because deciding what is a duplicate needs the whole
    # file: a line repeated within one export is two charges, the same line
    # arriving in a second export is one.
    to_create, decisions = _resolve_duplicates(candidates)

    # Belt to the guard above's braces: that one catches a line that arrived
    # misaligned, this one catches a line that arrived clean but disagrees with
    # a copy already on file. Both exist because a wrong row imports silently.
    near = _find_near_duplicates(
        [(line_number, txn) for line_number, txn in candidates
         if decisions[line_number] is None])

    for line_number, txn in candidates:
        skip_reason = decisions[line_number]
        if skip_reason:
            result.add(RowResult(line_number, RowResult.DUPLICATE, skip_reason))
        else:
            result.add(RowResult(line_number, RowResult.CREATED, transaction=txn,
                                 preview=previews[line_number],
                                 warning=_describe_near_duplicates(near.get(line_number))))
    result.rows.sort(key=lambda r: r.line_number)

    if not dry_run and to_create:
        with transaction.atomic():
            # bulk_create bypasses the immutability guard in save(), which is
            # correct here: these are all first-time inserts.
            WorkdayTransaction.objects.bulk_create(to_create)

    return result


# ---------------------------------------------------------------------------
# Staging an upload between "here is a file" and "yes, import it"
# ---------------------------------------------------------------------------

#: Where a not-yet-confirmed upload waits. Under the ordinary file store, so it
#: works the same on S3 as on a local disk.
STAGING_DIR = 'finance/staged_imports'

#: Anything left here longer than this was abandoned -- the Treasurer closed
#: the tab, or the preview said something they did not like.
STAGING_MAX_AGE = datetime.timedelta(hours=6)


def stage_upload(file_obj):
    """
    Park an upload where the confirmation step can pick it up again.

    The count in "you are about to add 253 lines" can only come from parsing
    the file, and the file is gone by the time the Treasurer answers: a browser
    will not re-submit an ``<input type=file>`` it never kept. So the bytes are
    written to the file store under an unguessable name and the parse runs
    twice -- once to count, once to import.

    :returns: the storage token to hand back on confirmation.
    """
    file_obj.seek(0)
    # The stored name is the token and nothing else. The Treasurer's own
    # filename travels in the session instead: it is theirs, it is displayed
    # back to them, and it has no business deciding a path in the file store.
    token = '%s-%s' % (timezone.now().strftime('%Y%m%d%H%M%S'), get_random_string(24))
    saved = default_storage.save('%s/%s' % (STAGING_DIR, token), ContentFile(file_obj.read()))
    file_obj.seek(0)
    return saved


def read_staged(token):
    """ The staged bytes as a file object, or ``None`` if it is no longer there. """
    if not token or not str(token).startswith(STAGING_DIR + '/'):
        # A token is a storage path we wrote. Anything else is someone trying
        # to make the importer read an arbitrary file out of the store.
        return None
    if not default_storage.exists(token):
        return None
    with default_storage.open(token, 'rb') as handle:
        return io.BytesIO(handle.read())


def discard_staged(token):
    """ Delete one staged upload. Safe to call twice. """
    if not token or not str(token).startswith(STAGING_DIR + '/'):
        return
    try:
        default_storage.delete(token)
    except (OSError, NotImplementedError):
        pass


def purge_stale_staged(now=None):
    """
    Delete staged uploads nobody came back for.

    Called on the way into a new upload rather than from a cron job: the only
    thing that creates these is an import, so the only moment one can have gone
    stale is when the next one starts.
    """
    now = now or timezone.now()
    try:
        _dirs, files = default_storage.listdir(STAGING_DIR)
    except (OSError, NotImplementedError, FileNotFoundError):
        return
    for name in files:
        path = '%s/%s' % (STAGING_DIR, name)
        try:
            modified = default_storage.get_modified_time(path)
        except (OSError, NotImplementedError):
            continue
        if timezone.is_naive(modified):
            modified = timezone.make_aware(modified, timezone.get_default_timezone())
        if now - modified > STAGING_MAX_AGE:
            discard_staged(path)
